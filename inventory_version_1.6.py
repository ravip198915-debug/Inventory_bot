import telebot
import sqlite3
import pytesseract
import cv2
import re
import pandas as pd
from datetime import datetime
from pyzbar.pyzbar import decode
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
import os
import schedule
import time
import threading

import shutil

# ================= BACKUP FUNCTION =================

def backup_database(send_to_telegram=False, chat_id=None):
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        backup_file = f"backup_{timestamp}.db"

        shutil.copy("shared_inventory.db", backup_file)

        print(f"✅ Backup created: {backup_file}")

        # 📤 Send to Telegram (optional)
        if send_to_telegram and chat_id:
            with open(backup_file, "rb") as f:
                bot.send_document(chat_id, f)

        return backup_file

    except Exception as e:
        print("Backup error:", e)
        return None


# ================= CLEAN OLD BACKUPS =================

def cleanup_backups():
    try:
        files = sorted([f for f in os.listdir() if f.startswith("backup_")])

        if len(files) > 7:
            for f in files[:-7]:
                os.remove(f)

        print("🧹 Old backups cleaned")

    except Exception as e:
        print("Backup cleanup error:", e)

# Tesseract path
#pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

BOT_TOKEN = BOT_TOKEN = "8298757534:AAGjgWyKpEz-9zDuNViZP1W6ggLSCSRlYWE"

bot = telebot.TeleBot(BOT_TOKEN)

#Create Permission Dictionary

ADMIN_IDS = [1412356698]

USER_PERMISSIONS = {

1412356698: ["admin"],

587636725: ["/start","/add_device","/stock","/report","/district_stock","/spare_avl","/pending_serials","/spare_balance","/cross_replacement","/district_report","/outward_spare","/inward_spare"],

7170102897: ["/start","/add_device","/stock","/report","/district_stock","/spare_avl","/pending_serials","/spare_balance","/cross_replacement","/district_report","/outward_spare","/inward_spare"],

8641112788: ["/start","/add_device","/stock","/report","/district_stock","/spare_avl","/pending_serials","/spare_balance","/cross_replacement","/district_report","/outward_spare","/inward_spare"],

8502188931: ["/start","/add_device","/stock","/report","/district_stock","/spare_avl","/pending_serials","/spare_balance","/cross_replacement","/district_report","/outward_spare","/inward_spare"],

7887580509: ["/start","/add_device","/stock","/report","/district_stock","/spare_avl","/pending_serials","/spare_balance","/cross_replacement","/district_report","/outward_spare","/inward_spare"],

838686002: ["/start","/add_device","/stock","/report","/district_stock","/spare_avl","/pending_serials","/spare_balance","/cross_replacement","/district_report","/outward_spare","/inward_spare"],

}

#Create Permission Check Function

def check_permission(message, command):

    user_id = message.from_user.id

    if user_id not in USER_PERMISSIONS:
        bot.send_message(message.chat.id,"❌ You are not authorized")
        return False

    if "admin" in USER_PERMISSIONS[user_id]:
        return True

    if command in USER_PERMISSIONS[user_id]:
        return True

    bot.send_message(message.chat.id,"❌ Permission denied for this command")
    return False

#dropdown
districts = [
"Salem",
"Dharmapuri",
"Namakkal",
"Krishnagiri",
"Villupuram",
"Kallakuruchi"
]

# DATABASE
conn = sqlite3.connect("shared_inventory.db", check_same_thread=False, timeout=60)

conn.execute("PRAGMA foreign_keys = ON")     # ✅ MUST
conn.execute("PRAGMA journal_mode=WAL")      # ✅ BETTER HERE

cursor = conn.cursor()

# ✅ INDEX FOR FAST MATCHING
cursor.execute("CREATE INDEX IF NOT EXISTS idx_devices_serial ON devices(serial_number)")
cursor.execute("CREATE INDEX IF NOT EXISTS idx_replacements_serial ON replacements(serial_number)")

conn.commit()

try:
    cursor.execute("ALTER TABLE spare_inward ADD COLUMN action TEXT")
except:
    pass

# Spare inward table
cursor.execute("""
CREATE TABLE IF NOT EXISTS spare_inward(
id INTEGER PRIMARY KEY AUTOINCREMENT,
district TEXT,
courier_name TEXT,
lr_number TEXT UNIQUE,
courier_date TEXT,
spare_name TEXT,
qty INTEGER,
remarks TEXT,
created_date TEXT
)
""")

# Device serial table
cursor.execute("""
CREATE TABLE IF NOT EXISTS devices(
id INTEGER PRIMARY KEY AUTOINCREMENT,
district TEXT,
courier_name TEXT,
lr_number TEXT,
courier_date TEXT,
model TEXT,
serial_number TEXT UNIQUE,
updated_date TEXT,
remarks TEXT
)
""")

# Replacement table (shared with spare_replace_bot)
cursor.execute("""
CREATE TABLE IF NOT EXISTS replacements(
id INTEGER PRIMARY KEY AUTOINCREMENT,
serial_number TEXT UNIQUE,
replaced_district TEXT,
replaced_date TEXT
)
""")

try:
    cursor.execute("ALTER TABLE replacements ADD COLUMN replaced_district TEXT")
except:
    pass

conn.commit()


# Opening Stock Table
cursor.execute("""
CREATE TABLE IF NOT EXISTS spare_opening_stock(
id INTEGER PRIMARY KEY AUTOINCREMENT,
district TEXT,
spare_name TEXT,
opening_qty INTEGER
)
""")

# 🔧 Remove duplicate district + spare rows
cursor.execute("""
DELETE FROM spare_opening_stock
WHERE id NOT IN (
    SELECT MIN(id)
    FROM spare_opening_stock
    GROUP BY district, spare_name
)
""")

# 🔒 Prevent future duplicates
cursor.execute("""
CREATE UNIQUE INDEX IF NOT EXISTS idx_district_spare
ON spare_opening_stock(district,spare_name)
""")

conn.commit()

#1️⃣ Create Outward Table

cursor.execute("""
CREATE TABLE IF NOT EXISTS spare_outward(
id INTEGER PRIMARY KEY AUTOINCREMENT,
dispatch_date TEXT,
courier_name TEXT,
lr_number TEXT,
spare_name TEXT,
qty INTEGER,
serial_number TEXT,
remarks TEXT,
created_date TEXT
)
""")

conn.commit()

try:
    cursor.execute("ALTER TABLE spare_outward ADD COLUMN district TEXT")
except:
    pass

# SYNC FUNCTION
def sync_replacements():

    cursor.execute("""
    UPDATE devices
    SET remarks = 'Replaced'
    WHERE TRIM(UPPER(serial_number)) IN (
        SELECT TRIM(UPPER(serial_number)) FROM replacements
    )
    AND (remarks IS NULL OR remarks != 'Replaced')
    """)

    conn.commit()
user_data = {}

# 👇 Stock Check Function

def get_available_stock(district, spare):

    cursor.execute("""
    SELECT
    o.opening_qty,
    COALESCE(SUM(CASE WHEN s.action='Receive' THEN s.qty END),0),
    COALESCE(SUM(CASE WHEN s.action='Send' THEN s.qty END),0)
    FROM spare_opening_stock o
    LEFT JOIN spare_inward s
    ON o.district = s.district
    AND o.spare_name = s.spare_name
    WHERE o.district=? AND o.spare_name=?
    GROUP BY o.opening_qty
    """, (district, spare))

    row = cursor.fetchone()

    if not row:
        return 0

    opening = row[0]
    received = row[1]
    sent = row[2]

    return opening + received - sent


# START
@bot.message_handler(commands=['start'])
def start(message):

    if message.from_user.id not in USER_PERMISSIONS:
        bot.send_message(message.chat.id,"❌ You are not authorized to use this bot")
        return

    bot.send_message(
        message.chat.id,
        "📦 Inventory Bot\n\n"

        "📥 Stock Entry\n"
        "/inward_spare - Add Spare Inward\n"
        "/add_device - Add Serial Number\n\n"

        "📊 Reports\n"
        "/report - Download Excel Report\n"
        "/lr_summary - LR Accuracy Report\n"
        "/pending_serials - Pending Serial Updates\n"
        "/district_report - District Stock Report\n\n"

        "📦 Inventory Status\n"
        "/stock - View Inventory\n"
        "/district_stock - District Inventory\n"
        "/courier_status - District Device Count\n"
        "/spare_avl - Spare Availability\n"
        "/spare_balance - Spare Balance\n\n"

        "🔄 Replacement Audit\n"
        "/cross_replacement - Cross District Replacement\n\n"

        "⚙️ Admin Commands\n"
        "/op - Add Opening Stock\n"
        "/edit_op - Edit Opening Stock\n"
    )
# INWARD SPARE
@bot.message_handler(commands=['inward_spare'])
def inward_spare(message):

    if not check_permission(message,"/inward_spare"):
        return

    user_data[message.chat.id] = {}

    markup = ReplyKeyboardMarkup(resize_keyboard=True)

    for d in districts:
        markup.add(KeyboardButton(d))

    msg = bot.send_message(message.chat.id,"Select District",reply_markup=markup)

    bot.register_next_step_handler(msg,inward_district)


def inward_district(message):

    user_data[message.chat.id]["district"] = message.text

    markup = ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add("Receive","Send")

    msg = bot.send_message(
        message.chat.id,
        "Select Option",
        reply_markup=markup
    )

    bot.register_next_step_handler(msg,inward_option)

def inward_option(message):

    user_data[message.chat.id]["action"] = message.text

    msg = bot.send_message(
        message.chat.id,
        "Enter Courier Name",
        reply_markup=ReplyKeyboardRemove()
    )

    bot.register_next_step_handler(msg,inward_courier)


def inward_courier(message):

    user_data[message.chat.id]["courier"] = message.text

    msg = bot.send_message(message.chat.id,"Enter LR Number")

    bot.register_next_step_handler(msg,inward_lr)


def inward_lr(message):

    user_data[message.chat.id]["lr"] = message.text

    msg = bot.send_message(message.chat.id,"Enter Courier Date (dd-mm-yyyy)")

    bot.register_next_step_handler(msg,inward_date)


def inward_date(message):

    user_data[message.chat.id]["date"] = message.text

    markup = ReplyKeyboardMarkup(resize_keyboard=True)

    markup.add("POS Device","Charger")
    markup.add("IRIS","Battery")
    markup.add("Others")

    msg = bot.send_message(message.chat.id,"Select Spare Type",reply_markup=markup)

    bot.register_next_step_handler(msg,inward_spare_type)


def inward_spare_type(message):

    if message.text == "Others":

        msg = bot.send_message(message.chat.id,"Enter Spare Name")

        bot.register_next_step_handler(msg,inward_spare_manual)

    else:

        user_data[message.chat.id]["spare"] = message.text

        msg = bot.send_message(message.chat.id,"Enter Quantity")

        bot.register_next_step_handler(msg,inward_qty)


def inward_spare_manual(message):

    user_data[message.chat.id]["spare"] = message.text

    msg = bot.send_message(message.chat.id,"Enter Quantity")

    bot.register_next_step_handler(msg,inward_qty)


def inward_qty(message):

    try:
        qty = int(message.text)
    except:
        bot.send_message(message.chat.id,"❌ Enter valid number for quantity")
        return

    user_data[message.chat.id]["qty"] = qty

    msg = bot.send_message(message.chat.id,"Enter Remarks")

    bot.register_next_step_handler(msg,inward_remarks)


def inward_remarks(message):

    data = user_data[message.chat.id]

    cursor.execute("""
    INSERT INTO spare_inward(
    district,courier_name,lr_number,courier_date,
    action,spare_name,qty,remarks,created_date
    )
    VALUES(?,?,?,?,?,?,?,?,?)
    """,(

    data["district"],
    data["courier"],
    data["lr"],
    data["date"],
    data["action"],
    data["spare"],
    data["qty"],
    message.text,
    datetime.now().strftime("%d-%m-%Y")

    ))

    conn.commit()

    bot.send_message(message.chat.id,"✅ Spare inward recorded")

# ADD DEVICE SERIAL
@bot.message_handler(commands=['add_device'])
def add_device(message):

    if not check_permission(message,"/add_device"):
        return

    msg = bot.send_message(message.chat.id,"Enter LR Number")

    bot.register_next_step_handler(msg,validate_lr)


def validate_lr(message):

    lr = message.text

    cursor.execute("""
    SELECT district,courier_name,courier_date,qty
    FROM spare_inward
    WHERE lr_number=?
    """,(lr,))

    row = cursor.fetchone()

    if not row:

        bot.send_message(message.chat.id,"❌ LR not found. Please enter inward first")
        return

    district = row[0]
    courier = row[1]
    date = row[2]
    inward_qty = row[3]

    cursor.execute("SELECT COUNT(*) FROM devices WHERE lr_number=?", (lr,))
    serial_count = cursor.fetchone()[0]

    if serial_count >= inward_qty:

        bot.send_message(message.chat.id,"⚠️ All serial numbers already updated")
        return

    user_data[message.chat.id] = {
        "lr": lr,
        "district": district,
        "courier": courier,
        "date": date
    }

    bot.send_message(
        message.chat.id,
        f"LR Found\nDistrict: {district}\nCourier: {courier}\nDate: {date}\n\nEnter Device Model"
    )

    bot.register_next_step_handler_by_chat_id(message.chat.id,get_model)

def get_model(message):

    user_data[message.chat.id]["model"] = message.text

    bot.send_message(message.chat.id,"📷 Send Serial Number Photo")

# SERIAL PHOTO
# SERIAL PHOTO
@bot.message_handler(content_types=['photo'])
def read_serial(message):

    # Detect mode
    outward_mode = False

    if message.chat.id in user_data and user_data[message.chat.id].get("mode") == "outward_photo":
        outward_mode = True

    # If not outward and no inward process started
    if not outward_mode and message.chat.id not in user_data:
        bot.send_message(message.chat.id,"Please start with /add_device first")
        return

    try:

        file_info = bot.get_file(message.photo[-1].file_id)
        downloaded_file = bot.download_file(file_info.file_path)

        file_path = f"temp_{message.chat.id}.jpg"

        with open(file_path,"wb") as f:
            f.write(downloaded_file)

        img = cv2.imread(file_path)

        serial = None

        # Barcode detection
        barcodes = decode(img)

        if barcodes:
            serial = barcodes[0].data.decode("utf-8")

        # OCR detection
        if not serial:

            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

            gray = cv2.GaussianBlur(gray,(5,5),0)

            thresh = cv2.threshold(gray,0,255,
            cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]

            text = pytesseract.image_to_string(gray)

            match = re.findall(r'[A-Z0-9]{15,20}', text)

            serial = match[0] if match else "NOT_FOUND"

        if serial == "NOT_FOUND":
            bot.send_message(
                message.chat.id,
                "❌ Serial not detected. Please resend clear image."
            )
            return

        serial = serial.strip().upper().replace(" ", "") 

        # DELETE TEMP FILE
        if os.path.exists(file_path):
            os.remove(file_path)

        # =========================
        # OUTWARD SERIAL SAVE
        # =========================

        if outward_mode:

            data = user_data[message.chat.id]

            # Prevent duplicate dispatch
            cursor.execute("SELECT * FROM spare_outward WHERE serial_number=?", (serial,))
            if cursor.fetchone():
                bot.send_message(message.chat.id,"⚠️ Serial already dispatched")
                return

            cursor.execute("""
            INSERT INTO spare_outward(
            dispatch_date,
            district,
            courier_name,
            lr_number,
            spare_name,
            qty,
            serial_number,
            remarks,
            created_date
            )
            VALUES(?,?,?,?,?,?,?,?,?)
            """,(

            data["date"],
            data["district"],
            data["courier"],
            data["lr"],
            data["spare"],
            1,
            serial,
            "",
            datetime.now().strftime("%d-%m-%Y")

            ))

            conn.commit()

            bot.send_message(message.chat.id,f"✅ Outward Serial recorded\n{serial}")

            return

        # =========================
        # INWARD DEVICE SAVE
        # =========================

        lr = user_data[message.chat.id]["lr"]

        cursor.execute("SELECT qty FROM spare_inward WHERE lr_number=?", (lr,))
        inward_qty = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM devices WHERE lr_number=?", (lr,))
        serial_count = cursor.fetchone()[0]

        if serial_count >= inward_qty:

            bot.send_message(message.chat.id,"⚠️ LR serial limit reached")
            return

        cursor.execute("""
        INSERT INTO devices(
        district,courier_name,lr_number,courier_date,
        model,serial_number,updated_date,remarks
        )
        VALUES(?,?,?,?,?,?,?,?)
        """,(

        user_data[message.chat.id].get("district",""),
        user_data[message.chat.id].get("courier",""),
        lr,
        user_data[message.chat.id].get("date",""),
        user_data[message.chat.id]["model"],
        serial,
        datetime.now().strftime("%d-%m-%Y"),
        "Available"

        ))

        conn.commit()

        bot.send_message(message.chat.id,f"✅ Serial added\n{serial}")

    except sqlite3.IntegrityError:

        bot.send_message(message.chat.id,"⚠️ Serial already exists")
# LR SUMMARY
@bot.message_handler(commands=['lr_summary'])
def lr_summary(message):

    if not check_permission(message,"/lr_summary"):
        return

    sync_replacements()

    cursor.execute("""
    SELECT 
    s.lr_number,
    s.spare_name,
    s.qty,
    COUNT(d.serial_number)
    FROM spare_inward s
    LEFT JOIN devices d
    ON s.lr_number = d.lr_number
    GROUP BY s.lr_number
    """)

    rows = cursor.fetchall()

    text = "📊 LR Accuracy Report\n\n"

    for r in rows:

        diff = r[2] - r[3]

        text += (
        f"LR : {r[0]}\n"
        f"Spare : {r[1]}\n"
        f"Received : {r[2]}\n"
        f"Serial Updated : {r[3]}\n"
        f"Difference : {diff}\n\n"
        )

    bot.send_message(message.chat.id,text)

# EXCEL REPORT
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

@bot.message_handler(commands=['report'])
def report(message):

    if not check_permission(message, "/report"):
        return

    sync_replacements()

    import os
    from datetime import datetime
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # ================= RAW DATA =================
    devices_df = pd.read_sql_query("SELECT * FROM devices", conn)
    spare_df = pd.read_sql_query("SELECT * FROM spare_inward", conn)
    outward_df = pd.read_sql_query("SELECT * FROM spare_outward", conn)

    # ================= POS SUMMARY (FIXED) =================
    pos_df = pd.read_sql_query("""
    SELECT 
        s.district,

        SUM(s.qty) AS received,

        SUM(
            (SELECT COUNT(*) FROM devices d WHERE d.lr_number = s.lr_number)
        ) AS serial_updated,

        SUM(s.qty) - SUM(
            (SELECT COUNT(*) FROM devices d WHERE d.lr_number = s.lr_number)
        ) AS pending,

        SUM(
            (SELECT COUNT(*) FROM devices d 
             WHERE d.lr_number = s.lr_number AND d.remarks='Replaced')
        ) AS replaced,

        SUM(
            (SELECT COUNT(*) FROM devices d 
             WHERE d.lr_number = s.lr_number AND d.remarks='Available')
        ) AS not_replaced

    FROM spare_inward s
    WHERE s.action='Receive' AND s.spare_name='POS Device'
    GROUP BY s.district
    ORDER BY s.district
    """, conn)

    # ================= IRIS SUMMARY (FIXED) =================
    iris_df = pd.read_sql_query("""
    SELECT 
        s.district,

        SUM(s.qty) AS received,

        SUM(
            (SELECT COUNT(*) FROM devices d WHERE d.lr_number = s.lr_number)
        ) AS serial_updated,

        SUM(s.qty) - SUM(
            (SELECT COUNT(*) FROM devices d WHERE d.lr_number = s.lr_number)
        ) AS pending,

        SUM(
            (SELECT COUNT(*) FROM devices d 
             WHERE d.lr_number = s.lr_number AND d.remarks='Replaced')
        ) AS replaced,

        SUM(
            (SELECT COUNT(*) FROM devices d 
             WHERE d.lr_number = s.lr_number AND d.remarks='Available')
        ) AS not_replaced

    FROM spare_inward s
    WHERE s.action='Receive' AND s.spare_name='IRIS'
    GROUP BY s.district
    ORDER BY s.district
    """, conn)

    # ================= COLUMN RENAME =================
    for df in [pos_df, iris_df]:
        df.rename(columns={"pending": "Serial Update Pending"}, inplace=True)

    # ================= ADD TOTAL =================
    def add_total(df):
        total = pd.DataFrame([{
            "district": "TOTAL",
            "received": df["received"].sum(),
            "serial_updated": df["serial_updated"].sum(),
            "Serial Update Pending": df["Serial Update Pending"].sum(),
            "replaced": df["replaced"].sum(),
            "not_replaced": df["not_replaced"].sum()
        }])
        return pd.concat([df, total], ignore_index=True)

    pos_df = add_total(pos_df)
    iris_df = add_total(iris_df)

    # ================= LR SUMMARY =================
    lr_df = pd.read_sql_query("""
    SELECT 
        s.district,
        s.lr_number,
        s.spare_name,
        s.qty AS received,
        COUNT(d.serial_number) AS serial_updated,
        (s.qty - COUNT(d.serial_number)) AS pending,
        SUM(CASE WHEN d.remarks='Replaced' THEN 1 ELSE 0 END) AS replaced,
        SUM(CASE WHEN d.remarks='Available' THEN 1 ELSE 0 END) AS not_replaced
    FROM spare_inward s
    LEFT JOIN devices d ON s.lr_number = d.lr_number
    WHERE s.action='Receive'
    GROUP BY s.district, s.lr_number, s.spare_name, s.qty
    """, conn)

    lr_df.rename(columns={"pending": "Serial Update Pending"}, inplace=True)

    # ================= SEND SUMMARY =================
    send_df = pd.read_sql_query("""
    SELECT 
        district,
        courier_name,
        lr_number,
        courier_date AS date,
        spare_name,
        qty
    FROM spare_inward
    WHERE action='Send'
    ORDER BY district
    """, conn)

    send_df.insert(0, "Sl.no", range(1, len(send_df) + 1))

    send_total = pd.DataFrame([{
        "Sl.no": "",
        "district": "TOTAL",
        "courier_name": "",
        "lr_number": "",
        "date": "",
        "spare_name": "",
        "qty": send_df["qty"].sum()
    }])

    send_df = pd.concat([send_df, send_total], ignore_index=True)

    # ================= DISTRICT SPARE =================
    # ================= DISTRICT SPARE =================
    # ================= DISTRICT SPARE =================
    district_df = pd.read_sql_query("""
    WITH inward AS (
        SELECT 
            district, 
            spare_name,
            SUM(CASE WHEN action='Receive' THEN qty ELSE 0 END) AS received_qty,
            SUM(CASE WHEN action='Send' THEN qty ELSE 0 END) AS sent_qty
        FROM spare_inward
        GROUP BY district, spare_name
    ),

    replacements_agg AS (
        SELECT 
            r.replaced_district AS district,

            CASE 
                WHEN LOWER(model) LIKE '%pos%' THEN 'POS Device'
                WHEN LOWER(model) LIKE '%iris%' THEN 'IRIS'
                WHEN LOWER(model) LIKE '%bio%' THEN 'Biometric'
                WHEN LOWER(model) LIKE '%bat%' THEN 'Battery'
                WHEN LOWER(model) LIKE '%char%' THEN 'Charger'
                WHEN LOWER(model) LIKE '%scan%' THEN 'Scanning Glass'
                ELSE TRIM(model) 
            END AS spare_name,

            COUNT(*) AS replaced 

        FROM replacements r

        LEFT JOIN devices d 
            ON r.serial_number = d.serial_number

        GROUP BY r.replaced_district, spare_name
    )

    SELECT 
        o.district,
        o.spare_name,
        o.opening_qty,

        COALESCE(i.received_qty,0) AS received_qty,
        COALESCE(i.sent_qty,0) AS sent_qty,

        o.opening_qty + 
        COALESCE(i.received_qty,0) - 
        COALESCE(i.sent_qty,0) AS balance_qty,

        COALESCE(r.replaced,0) AS replaced,
        
        (o.opening_qty + COALESCE(i.received_qty,0) - COALESCE(r.replaced,0)) AS not_replaced,

        (o.opening_qty + COALESCE(i.received_qty,0) - COALESCE(r.replaced,0)) AS faulty_in_district 


    FROM spare_opening_stock o

    LEFT JOIN inward i 
        ON o.district = i.district 
        AND o.spare_name = i.spare_name

    LEFT JOIN replacements_agg r  
        ON LOWER(TRIM(o.district)) = LOWER(TRIM(r.district))
        AND LOWER(TRIM(o.spare_name)) = LOWER(TRIM(r.spare_name))

    ORDER BY o.district, o.spare_name
    """, conn)


# ================= FILE =================
    file = f"Inventory_Report_{datetime.now().strftime('%d-%b-%Y_%H-%M')}.xlsx"

    with pd.ExcelWriter(file, engine="openpyxl") as writer:
        devices_df.to_excel(writer, sheet_name="Devices", index=False)
        spare_df.to_excel(writer, sheet_name="Spare_Inward", index=False)
        outward_df.to_excel(writer, sheet_name="Outward", index=False)

        pos_df.to_excel(writer, sheet_name="POS_Summary", index=False)
        iris_df.to_excel(writer, sheet_name="IRIS_Summary", index=False)
        lr_df.to_excel(writer, sheet_name="LR_Summary", index=False)
        send_df.to_excel(writer, sheet_name="SEND_Summary", index=False)
        district_df.to_excel(writer, sheet_name="District_Spare_Balance", index=False)

    # ================= FORMATTING =================
    wb = load_workbook(file)

    header_fill = PatternFill(start_color="305496", fill_type="solid")
    total_fill = PatternFill(start_color="FFD966", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", fill_type="solid")

    header_font = Font(bold=True, color="FFFFFF")
    total_font = Font(bold=True)
    red_font = Font(color="FFFFFF", bold=True)

    align = Alignment(horizontal="center", vertical="center")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.freeze_panes = "A2"

        # Header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align
            cell.border = border

    # ✅ FIND FAULTY COLUMN (MISSING IN YOUR CODE)
        faulty_col = None
        for idx, cell in enumerate(ws[1]):
            if cell.value in ["faulty_in_district", "Faulty in District"]:
                faulty_col = idx
                break
        
        # Rows
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = align
                cell.border = border

            if str(row[0].value) == "TOTAL":
                for cell in row:
                    cell.fill = total_fill
                    cell.font = total_font

        # 🔴 Faulty >= 10 highlight
            if faulty_col is not None:
                val = row[faulty_col].value
                if isinstance(val, (int, float)) and val >= 10:
                    for cell in row:
                        cell.fill = red_fill
                        cell.font = red_font


        # Auto width
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = max_len + 3

    wb.save(file)

    with open(file, "rb") as f:
        bot.send_document(message.chat.id, f)

    os.remove(file)

#1️⃣ /stock — View All Inventory

@bot.message_handler(commands=['stock'])
def stock(message):

    if not check_permission(message,"/stock"):
        return

    sync_replacements()

    cursor.execute("""
    SELECT 
        COUNT(*),
        SUM(CASE WHEN remarks='Available' THEN 1 ELSE 0 END),
        SUM(CASE WHEN remarks='Replaced' THEN 1 ELSE 0 END)
    FROM devices
    """)

    total, available, replaced = cursor.fetchone()

    # Handle None values
    available = available or 0
    replaced = replaced or 0

    text = f"""
📦 Inventory Summary

Total Devices : {total}
Available     : {available}
Replaced      : {replaced}
"""

    bot.send_message(message.chat.id, text)

#2️⃣ /delete SERIAL — Delete Serial

@bot.message_handler(commands=['delete'])
def delete_serial(message):

    if not check_permission(message,"/delete"):
        return

    parts = message.text.split()

    if len(parts) < 2:
        bot.send_message(message.chat.id,"Usage:\n/delete SERIAL")
        return

    serial = parts[1]

    cursor.execute("SELECT * FROM devices WHERE serial_number=?", (serial,))
    row = cursor.fetchone()

    if not row:
        bot.send_message(message.chat.id,"❌ Serial not found")
        return

    cursor.execute("DELETE FROM devices WHERE serial_number=?", (serial,))
    conn.commit()

    bot.send_message(message.chat.id,f"✅ Serial {serial} deleted")

#3️⃣ /courier_status — District Device Count

@bot.message_handler(commands=['courier_status'])
def courier_status(message):

    if not check_permission(message,"/courier_status"):
        return

    sync_replacements()

    cursor.execute("""
    SELECT district, COUNT(*)
    FROM devices
    GROUP BY district
    """)

    rows = cursor.fetchall()

    if not rows:
        bot.send_message(message.chat.id,"No inventory data")
        return

    text = "📦 District Courier Status\n\n"

    for r in rows:
        text += f"{r[0]} : {r[1]} devices\n"

    bot.send_message(message.chat.id,text)

#4️⃣ /district_stock — District Inventory Summary

@bot.message_handler(commands=['district_stock'])
def district_stock(message):

    if not check_permission(message,"/district_stock"):
        return

    sync_replacements()

    cursor.execute("""
    SELECT district, COUNT(*)
    FROM devices
    GROUP BY district
    """)

    rows = cursor.fetchall()

    if not rows:
        bot.send_message(message.chat.id,"No inventory data")
        return

    text = "📦 District Inventory Stock\n\n"

    total = 0

    for r in rows:
        text += f"{r[0]} : {r[1]}\n"
        total += r[1]

    text += f"\nTotal Devices : {total}"

    bot.send_message(message.chat.id,text)

#5️⃣ /spare_avl — Spare Availability Summary

@bot.message_handler(commands=['spare_avl'])
def spare_avl(message):

    if not check_permission(message,"/spare_avl"):
        return

    sync_replacements()

    cursor.execute("""
    SELECT 
    district,
    SUM(CASE WHEN remarks='Available' THEN 1 ELSE 0 END),
    SUM(CASE WHEN remarks='Replaced' THEN 1 ELSE 0 END),
    COUNT(*)
    FROM devices
    GROUP BY district
    ORDER BY district
    """)

    rows = cursor.fetchall()

    if not rows:
        bot.send_message(message.chat.id,"No inventory data")
        return

    text = "📊 Spare Availability Summary\n\n"

    total_available = 0
    total_replaced = 0

    for r in rows:

        district = r[0]
        available = r[1]
        replaced = r[2]
        total = r[3]

        text += f"{district}\nAvailable : {available}\nReplaced : {replaced}\nTotal : {total}\n\n"

        total_available += available
        total_replaced += replaced

    text += f"Overall Available : {total_available}\nOverall Replaced : {total_replaced}"

    bot.send_message(message.chat.id,text)

#/pending_serials Command

@bot.message_handler(commands=['pending_serials'])
def pending_serials(message):

    if not check_permission(message,"/pending_serials"):
        return

    sync_replacements()

    cursor.execute("""
    SELECT 
    s.lr_number,
    s.spare_name,
    s.qty,
    COUNT(d.serial_number)
    FROM spare_inward s
    LEFT JOIN devices d
    ON s.lr_number = d.lr_number
    GROUP BY s.lr_number
    """)

    rows = cursor.fetchall()

    text = "⚠️ Pending Serial Updates\n\n"

    found = False

    for r in rows:

        lr = r[0]
        spare = r[1]
        received = r[2]
        updated = r[3]

        pending = received - updated

        if pending > 0:

            found = True

            text += (
            f"LR : {lr}\n"
            f"Spare : {spare}\n"
            f"Received : {received}\n"
            f"Serial Updated : {updated}\n"
            f"Pending : {pending}\n\n"
            )

    if not found:
        text = "✅ No pending serial updates"

    bot.send_message(message.chat.id,text)



#2️⃣ Spare selection (handle Others)

def op_spare(message):

    if message.text == "Others":

        msg = bot.send_message(
            message.chat.id,
            "Enter Spare Name",
            reply_markup=ReplyKeyboardRemove()
        )

        bot.register_next_step_handler(msg, op_spare_manual)

    else:

        user_data[message.chat.id]["spare"] = message.text

        op_select_district(message)

#3️⃣ Manual spare name

def op_spare_manual(message):

    user_data[message.chat.id]["spare"] = message.text

    op_select_district(message)

#4️⃣ District selection

def op_select_district(message):

    markup = ReplyKeyboardMarkup(resize_keyboard=True)

    for d in districts:
        markup.add(KeyboardButton(d))

    msg = bot.send_message(
        message.chat.id,
        "Select District",
        reply_markup=markup
    )

    bot.register_next_step_handler(msg, op_district)

#5️⃣ District → Quantity

def op_district(message):

    district = message.text

    if district not in districts:
        bot.send_message(message.chat.id,"❌ Please select district from list")
        return

    user_data[message.chat.id]["district"] = district

    msg = bot.send_message(
        message.chat.id,
        "Enter Opening Quantity",
        reply_markup=ReplyKeyboardRemove()
    )

    bot.register_next_step_handler(msg, op_qty)

#6️⃣ Save opening stock

def op_qty(message):

    data = user_data[message.chat.id]

    try:
        qty = int(message.text)
    except:
        bot.send_message(message.chat.id,"❌ Quantity must be greater than 0")
        return

    district = data["district"]
    spare = data["spare"]

    cursor.execute("""
    SELECT opening_qty FROM spare_opening_stock
    WHERE district=? AND spare_name=?
    """,(district,spare))

    row = cursor.fetchone()

    if row:

        current_qty = row[0]
        new_qty = current_qty + qty

        cursor.execute("""
        UPDATE spare_opening_stock
        SET opening_qty=?
        WHERE district=? AND spare_name=?
        """,(new_qty,district,spare))

        msg = "➕ Stock Added"

    else:

        new_qty = qty

        cursor.execute("""
        INSERT INTO spare_opening_stock(district,spare_name,opening_qty)
        VALUES(?,?,?)
        """,(district,spare,new_qty))

        msg = "✅ Opening Stock Created"

    conn.commit()

    bot.send_message(
        message.chat.id,
        f"{msg}\n\n"
        f"District : {district}\n"
        f"Spare : {spare}\n"
        f"Added Qty : {qty}\n"
        f"Total Qty : {new_qty}"
    )
#2️⃣ Restrict /op command

@bot.message_handler(commands=['op'])
def opening_stock(message):

    if message.from_user.id not in ADMIN_IDS:
        bot.send_message(message.chat.id,"❌ Admin only command")
        return

    user_data[message.chat.id] = {}

    markup = ReplyKeyboardMarkup(resize_keyboard=True)

    markup.add("POS Device","Charger")
    markup.add("IRIS","Battery")
    markup.add("Others")

    msg = bot.send_message(message.chat.id,"Select Spare",reply_markup=markup)

    bot.register_next_step_handler(msg, op_spare)


#3️⃣ Restrict /edit_op command

@bot.message_handler(commands=['edit_op'])
def edit_opening_stock(message):

    if message.from_user.id not in ADMIN_IDS:
        bot.send_message(message.chat.id,"❌ Admin only command")
        return

    user_data[message.chat.id] = {}

    markup = ReplyKeyboardMarkup(resize_keyboard=True)

    markup.add("POS Device","Charger")
    markup.add("IRIS","Battery")
    markup.add("Others")

    msg = bot.send_message(
        message.chat.id,
        "Select Spare to Edit",
        reply_markup=markup
    )

    bot.register_next_step_handler(msg, edit_op_spare)

#1️⃣ Handle Spare Selection

def edit_op_spare(message):

    if message.text == "Others":

        msg = bot.send_message(
            message.chat.id,
            "Enter Spare Name",
            reply_markup=ReplyKeyboardRemove()
        )

        bot.register_next_step_handler(msg, edit_op_spare_manual)

    else:

        user_data[message.chat.id]["spare"] = message.text

        edit_op_district(message)

#2️⃣ Manual Spare Name

def edit_op_spare_manual(message):

    user_data[message.chat.id]["spare"] = message.text

    edit_op_district(message)

#3️⃣ Select District

def edit_op_district(message):

    markup = ReplyKeyboardMarkup(resize_keyboard=True)

    for d in districts:
        markup.add(KeyboardButton(d))

    msg = bot.send_message(
        message.chat.id,
        "Select District",
        reply_markup=markup
    )

    bot.register_next_step_handler(msg, edit_op_qty)


#4️⃣ Enter New Quantity

def edit_op_qty(message):

    data = user_data[message.chat.id]

    district = message.text

    msg = bot.send_message(
        message.chat.id,
        "Enter New Opening Quantity",
        reply_markup=ReplyKeyboardRemove()
    )

    user_data[message.chat.id]["district"] = district

    bot.register_next_step_handler(msg, update_op_qty)


#5️⃣ Update Database

def update_op_qty(message):

    data = user_data[message.chat.id]

    try:
        qty = int(message.text)
    except:
        bot.send_message(message.chat.id,"❌ Enter valid quantity")
        return

    # 🔍 Check if opening stock exists
    cursor.execute(
        "SELECT * FROM spare_opening_stock WHERE district=? AND spare_name=?",
        (data["district"], data["spare"])
    )

    row = cursor.fetchone()

    if not row:
        bot.send_message(message.chat.id,"❌ Opening stock not found")
        return

    # ✅ Update quantity
    cursor.execute("""
    UPDATE spare_opening_stock
    SET opening_qty=?
    WHERE district=? AND spare_name=?
    """,(
        qty,
        data["district"],
        data["spare"]
    ))

    conn.commit()

    bot.send_message(
        message.chat.id,
        f"✅ Opening Stock Updated\n\nDistrict : {data['district']}\nSpare : {data['spare']}\nNew Qty : {qty}"
    )
#/spare_balance Command

@bot.message_handler(commands=['spare_balance'])
def spare_balance(message):

    if not check_permission(message,"/spare_balance"):
        return

    cursor.execute("""
    SELECT
    o.district,
    o.spare_name,
    o.opening_qty,
    
    COALESCE(SUM(CASE WHEN s.action='Receive' THEN s.qty END),0) AS received_qty,
    
    COALESCE(SUM(CASE WHEN s.action='Send' THEN s.qty END),0) AS sent_qty

    FROM spare_opening_stock o

    LEFT JOIN spare_inward s
    ON o.district = s.district
    AND o.spare_name = s.spare_name

    GROUP BY o.district,o.spare_name,o.opening_qty
    ORDER BY o.district
    """)

    rows = cursor.fetchall()

    if not rows:
        bot.send_message(message.chat.id,"No spare stock data")
        return

    text = "📦 District Spare Balance\n\n"

    current_district = ""

    for r in rows:

        district = r[0]
        spare = r[1]
        opening = r[2]
        received = r[3]
        sent = r[4]

        balance = opening + received - sent

        if district != current_district:
            text += f"\n🏢 {district}\n"
            current_district = district

        text += f"{spare} : {balance}\n"

    bot.send_message(message.chat.id,text)

#2️⃣ Excel Query for Cross District Replacement

@bot.message_handler(commands=['cross_replacement'])
def cross_replacement(message):

    if not check_permission(message,"/cross_replacement"):
        return

    query = """
    SELECT
    d.serial_number,
    d.model,
    d.district AS updated_district,
    r.replaced_district,
    d.updated_date,
    r.replaced_date
    FROM devices d
    JOIN replacements r
    ON d.serial_number = r.serial_number
    WHERE d.district != r.replaced_district
    """

    df = pd.read_sql_query(query, conn)

    if df.empty:
        bot.send_message(message.chat.id,"✅ No cross district replacements")
        return

    file = "cross_district_replacements.xlsx"

    df.to_excel(file,index=False)

    with open(file,"rb") as f:
        bot.send_document(message.chat.id,f)

    if os.path.exists(file):
        os.remove(file)

#district_report

@bot.message_handler(commands=['district_report'])
def district_report(message):

    if not check_permission(message,"/district_report"):
        return

    bot.send_message(message.chat.id,"Generating district report...")

    query = """
    SELECT
    o.district,
    o.spare_name,
    o.opening_qty,

    COALESCE(SUM(CASE WHEN s.action='Receive' THEN s.qty END),0) AS received_qty,
    COALESCE(SUM(CASE WHEN s.action='Send' THEN s.qty END),0) AS sent_qty,

    o.opening_qty +
    COALESCE(SUM(CASE WHEN s.action='Receive' THEN s.qty END),0) -
    COALESCE(SUM(CASE WHEN s.action='Send' THEN s.qty END),0) AS balance

    FROM spare_opening_stock o
    LEFT JOIN spare_inward s
    ON o.district = s.district
    AND o.spare_name = s.spare_name

    GROUP BY o.district,o.spare_name,o.opening_qty
    ORDER BY o.district
    """

    df = pd.read_sql_query(query, conn)

    if df.empty:
        bot.send_message(message.chat.id,"No district stock data")
        return

    file = "district_stock_report.xlsx"

    df.to_excel(file,index=False)

    with open(file,"rb") as f:
        bot.send_document(message.chat.id,f)

    if os.path.exists(file):
        os.remove(file)

#Add users through telegram

# ADD USER (Admin Only)

@bot.message_handler(commands=['add_user'])
def add_user(message):

    if message.from_user.id not in ADMIN_IDS:
        bot.send_message(message.chat.id,"❌ Admin only command")
        return

    msg = bot.send_message(
        message.chat.id,
        "Send User ID to add"
    )

    bot.register_next_step_handler(msg,save_user)


def save_user(message):

    try:
        new_user = int(message.text)
    except:
        bot.send_message(message.chat.id,"❌ Invalid User ID")
        return

    USER_PERMISSIONS[new_user] = [
        "/start",
        "/inward_spare",
        "/add_device",
        "/stock",

        "/report",
        "/district_stock",
        "/spare_avl",
        "/pending_serials",
        "/spare_balance",
        "/cross_replacement",
        "/district_report"
    ]

    bot.send_message(message.chat.id,f"✅ User {new_user} added successfully")


#Add bot restart command (Admin)

@bot.message_handler(commands=['restart'])
def restart_bot(message):

    if message.from_user.id not in ADMIN_IDS:
        return

    bot.send_message(message.chat.id,"♻️ Restarting bot...")
    os._exit(0)


#all commands

# SHOW ALL COMMANDS

# SHOW ALL COMMANDS

@bot.message_handler(commands=['commands'])
def show_commands(message):

    if not check_permission(message,"/start"):
        return

    bot.send_message(
        message.chat.id,
        "📦 Inventory Bot Commands\n\n"

        "📥 Stock Entry\n"
        "/inward_spare\n"
        "/add_device\n\n"

        "📊 Reports\n"
        "/report\n"
        "/lr_summary\n"
        "/pending_serials\n"
        "/district_report\n\n"

        "📦 Inventory Status\n"
        "/stock\n"
        "/district_stock\n"
        "/courier_status\n"
        "/spare_avl\n"
        "/spare_balance\n\n"

        "🔄 Replacement Audit\n"
        "/cross_replacement\n\n"

        "⚙️ Admin Commands\n"
        "/op\n"
        "/edit_op\n"
        "/add_user\n"
        "/bot_status\n\n"

        "🗑 Data Management\n"
        "/delete SERIAL\n"
    )
#/bot_status admin command to monitor your system health

# BOT STATUS (Admin Only)

@bot.message_handler(commands=['bot_status'])
def bot_status(message):

    if message.from_user.id not in ADMIN_IDS:
        bot.send_message(message.chat.id,"❌ Admin only command")
        return

    try:

        # Database size
        db_size = os.path.getsize("shared_inventory.db") / (1024 * 1024)
        db_size = round(db_size,2)

        # Total devices
        cursor.execute("SELECT COUNT(*) FROM devices")
        total_devices = cursor.fetchone()[0]

        # Total LR
        cursor.execute("SELECT COUNT(*) FROM spare_inward")
        total_lr = cursor.fetchone()[0]

        # Total spare inward qty
        cursor.execute("SELECT SUM(qty) FROM spare_inward")
        total_spares = cursor.fetchone()[0]
        if total_spares is None:
            total_spares = 0

        # Total users
        total_users = len(USER_PERMISSIONS)

        text = (
            "📊 BOT SYSTEM STATUS\n\n"
            f"🗄 Database Size : {db_size} MB\n"
            f"📦 Total Devices : {total_devices}\n"
            f"🚚 Total LR Entries : {total_lr}\n"
            f"🔧 Total Spares Qty : {total_spares}\n"
            f"👤 Total Users : {total_users}\n"
            f"🟢 Bot Status : Running"
        )

        bot.send_message(message.chat.id,text)

    except Exception as e:

        bot.send_message(message.chat.id,f"⚠️ Error reading system status\n{e}")


#2️⃣ Command /outward_spare

@bot.message_handler(commands=['outward_spare'])
def outward_spare(message):

    if not check_permission(message,"/outward_spare"):
        return

    user_data[message.chat.id] = {}

    markup = ReplyKeyboardMarkup(resize_keyboard=True)

    for d in districts:
        markup.add(KeyboardButton(d))

    msg = bot.send_message(message.chat.id, "Select District", reply_markup=markup)

    bot.register_next_step_handler(msg, outward_district)

def outward_district(message):

    user_data[message.chat.id]["district"] = message.text

    msg = bot.send_message(
        message.chat.id,
        "Enter Dispatch Date (dd-mm-yyyy)",
        reply_markup=ReplyKeyboardRemove()
    )

    bot.register_next_step_handler(msg, outward_date)



#3️⃣ Courier & LR

def outward_date(message):

    user_data[message.chat.id]["date"] = message.text

    msg = bot.send_message(message.chat.id,"Enter Courier Name")

    bot.register_next_step_handler(msg,outward_courier)


def outward_courier(message):

    user_data[message.chat.id]["courier"] = message.text

    msg = bot.send_message(message.chat.id,"Enter LR Number")

    bot.register_next_step_handler(msg,outward_lr)

#4️⃣ Spare Dropdown

def outward_lr(message):

    user_data[message.chat.id]["lr"] = message.text

    markup = ReplyKeyboardMarkup(resize_keyboard=True)

    markup.add("POS Device","Charger")
    markup.add("IRIS","Battery")
    markup.add("Others")

    msg = bot.send_message(message.chat.id,"Select Spare Type",reply_markup=markup)

    bot.register_next_step_handler(msg,outward_spare_type)

#5️⃣ Spare Selection

def outward_spare_type(message):

    if message.text == "Others":

        msg = bot.send_message(message.chat.id, "Enter Spare Name")
        bot.register_next_step_handler(msg, outward_spare_manual)

    else:

        user_data[message.chat.id]["spare"] = message.text

        msg = bot.send_message(message.chat.id, "Enter Quantity")
        bot.register_next_step_handler(msg, outward_qty)


# ================= QUANTITY FUNCTION =================

def outward_qty(message):

    try:
        qty = int(message.text)
    except:
        bot.send_message(message.chat.id, "❌ Enter valid quantity")
        return

    data = user_data[message.chat.id]

    district = data["district"]
    spare = data["spare"]

    # ✅ STOCK VALIDATION
    available = get_available_stock(district, spare)

    if qty > available:
        bot.send_message(
            message.chat.id,
            f"❌ Not enough stock\n\nAvailable: {available}\nTrying to send: {qty}"
        )
        return

    # ✅ SAVE QTY
    user_data[message.chat.id]["qty"] = qty
    user_data[message.chat.id]["serial_count"] = 0

    if spare == "POS Device":

        ask_serial_method(message)

    else:

        cursor.execute("""
        INSERT INTO spare_outward(
        dispatch_date,
        district,
        courier_name,
        lr_number,
        spare_name,
        qty,
        serial_number,
        remarks,
        created_date
        )
        VALUES(?,?,?,?,?,?,?,?,?)
        """, (

        data["date"],
        data["district"],   # ✅ IMPORTANT
        data["courier"],
        data["lr"],
        data["spare"],
        qty,
        "",
        "",
        datetime.now().strftime("%d-%m-%Y")

        ))

        conn.commit()

        bot.send_message(message.chat.id, "✅ Outward recorded successfully")

        # ✅ CLEAN USER DATA
        del user_data[message.chat.id]


#6️⃣ Serial Capture Method

def ask_serial_method(message):

    markup = ReplyKeyboardMarkup(resize_keyboard=True)

    markup.add("📷 Photo","✍️ Manual Entry")
    markup.add("📄 Excel Upload")

    msg = bot.send_message(
        message.chat.id,
        "Select Serial Entry Method",
        reply_markup=markup
    )

    bot.register_next_step_handler(msg,serial_method)

#7️⃣ Manual Serial Entry

def serial_method(message):

    if message.text == "✍️ Manual Entry":

        msg = bot.send_message(message.chat.id,"Enter Serial Number")

        bot.register_next_step_handler(msg,save_manual_serial)

    elif message.text == "📷 Photo":

         user_data[message.chat.id]["mode"] = "outward_photo" 

         bot.send_message(message.chat.id,"Send Serial Photo")

    elif message.text == "📄 Excel Upload":

        bot.send_message(message.chat.id,"Upload Excel with serial numbers")

#8️⃣ Save Manual Serial

def save_manual_serial(message):

    serial = message.text.strip().upper().replace(" ", "")

    data = user_data[message.chat.id]

    # ✅ STOCK CHECK (prevents extra serial entry)
    available = get_available_stock(data["district"], data["spare"])

    remaining = data["qty"] - data["serial_count"]

    if available < 1 or remaining <= 0:
        bot.send_message(message.chat.id, "❌ No stock available")
        return

    if data["serial_count"] >= data["qty"]:
        bot.send_message(message.chat.id, "⚠️ Serial limit reached")
        return

    # 🔒 Duplicate serial protection (already dispatched)
    cursor.execute("SELECT * FROM spare_outward WHERE serial_number=?", (serial,))
    if cursor.fetchone():
        bot.send_message(message.chat.id, "⚠ Serial already dispatched")
        return

    # 🔎 Check if serial exists in inventory
    cursor.execute("SELECT * FROM devices WHERE serial_number=?", (serial,))
    if not cursor.fetchone():
        bot.send_message(message.chat.id, "⚠ Serial not found in inventory")
        return

    # ✅ INSERT INTO OUTWARD (with district)
    cursor.execute("""
    INSERT INTO spare_outward(
    dispatch_date,
    district,
    courier_name,
    lr_number,
    spare_name,
    qty,
    serial_number,
    remarks,
    created_date
    )
    VALUES(?,?,?,?,?,?,?,?,?)
    """,
    (

    data["date"],
    data["district"],   # ✅ IMPORTANT
    data["courier"],
    data["lr"],
    data["spare"],
    1,
    serial,
    "",
    datetime.now().strftime("%d-%m-%Y")

    ))

    conn.commit()

    data["serial_count"] += 1

    # 🔁 Continue until required qty reached
    if data["serial_count"] < data["qty"]:

        msg = bot.send_message(
            message.chat.id,
            f"Enter Serial Number ({data['serial_count']+1}/{data['qty']})"
        )

        bot.register_next_step_handler(msg, save_manual_serial)

    else:

        bot.send_message(message.chat.id, "✅ All serial numbers recorded")

        # ✅ CLEAN SESSION
        del user_data[message.chat.id]

def send_large_message(chat_id, text):
    chunk_size = 4000
    for i in range(0, len(text), chunk_size):
        bot.send_message(chat_id, text[i:i+chunk_size])



#/cancel Command

@bot.message_handler(commands=['cancel'])
def cancel_process(message):

    chat_id = message.chat.id

    if chat_id in user_data:
        del user_data[chat_id]

    bot.clear_step_handler_by_chat_id(chat_id)

    bot.send_message(
        chat_id,
        "❌ Process cancelled.",
        reply_markup=ReplyKeyboardRemove()
    )


#✅ Add Auto Cleanup Function

# AUTO CLEAN OLD DATA (KEEP LAST 30 DAYS)

def clean_old_data():
    try:
        cursor.execute(
            "DELETE FROM spare_inward "
            "WHERE date(substr(created_date,7,4) || '-' || substr(created_date,4,2) || '-' || substr(created_date,1,2)) "
            "< date('now','-90 day')"
        )

        cursor.execute(
            "DELETE FROM devices "
            "WHERE date(substr(updated_date,7,4) || '-' || substr(updated_date,4,2) || '-' || substr(updated_date,1,2)) "
            "< date('now','-90 day')"
        )

        cursor.execute(
            "DELETE FROM replacements "
            "WHERE date(substr(replaced_date,7,4) || '-' || substr(replaced_date,4,2) || '-' || substr(replaced_date,1,2)) "
            "< date('now','-90 day')"
        )

        conn.commit()

        cursor.execute("VACUUM")

        print("🧹 Old records cleaned and database optimized")

    except Exception as e:
        print("Error in cleanup:", e)
#4️⃣ Create Scheduler

# DAILY CLEANUP SCHEDULER

def run_scheduler():

    schedule.every().day.at("06:30").do(backup_database)   # ✅ Backup first
    schedule.every().day.at("06:45").do(clean_old_data)    # ✅ Then cleanup DB
    schedule.every().day.at("07:00").do(cleanup_backups)   # ✅ Cleanup backups

    while True:
        try:
            schedule.run_pending()
            time.sleep(60)
        except Exception as e:
            print("Scheduler error:", e)


print("Inventory Bot Running...")

def auto_sync_loop():
    while True:
        try:
            sync_replacements()
            time.sleep(60)  # every 1 min
        except Exception as e:
            print("Sync error:", e)

threading.Thread(target=auto_sync_loop, daemon=True).start()

# Start cleanup scheduler
threading.Thread(target=run_scheduler, daemon=True).start()

while True:
    try:
        bot.infinity_polling(timeout=60, long_polling_timeout=60)
    except Exception as e:
        print("Bot error:", e)
        time.sleep(10)	